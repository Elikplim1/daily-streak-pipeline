"""
Accumulator Agent — builds a daily evidence-backed selection betslip.

Scans flagged_opportunities for today's scan, ranks opportunities by a
composite evidence score, selects the best MAX_SELECTIONS with
diversification constraints, and sends a Telegram message + Excel report
with reasoning for each selection.

This module ONLY selects and reports. It never modifies signals.
Principle: "Evidence is immutable. Decisions are policy."
"""
import logging
import os
from datetime import datetime
from typing import List, Optional, Tuple
from collections import defaultdict
from dataclasses import dataclass, field

from src.db import get_cursor
from src.config import SHADOW_MODE, STREAK_WINDOW, MODERATE_SIGNAL_MIN
from src.telegram_notifier import send_telegram_message

logger = logging.getLogger(__name__)


# ─── Configuration ───────────────────────────────────────────────────

# Leagues validated by the Phase 1 backtest (scripts/backtest_full.py) —
# 50+ completed fixtures in BOTH season 2024 and season 2025. There is no
# TIER_1/TIER_2 split here: every league below was backtested identically.
# (A draft spec proposed a "TIER_1 — confirmed strong home advantage" list
# of Nordic/Baltic leagues plus MLS/K League 1/Danish Superliga/Swedish
# Allsvenskan for TIER_2. Dropped: those leagues have zero season-2024
# fixture data in this database, so no backtest run in this codebase could
# have validated them — shipping that split would have broadcast false
# "proven" claims in real selection reasoning.)
VALIDATED_LEAGUES = [
    '2. Bundesliga', 'Austrian Bundesliga', 'Belgian Pro League', 'Bundesliga',
    'Championship', 'Eerste Divisie', 'Eredivisie', 'La Liga', 'League One',
    'League Two', 'Liga Portugal', 'Ligue 1', 'Ligue 2', 'Premier League',
    'Saudi Pro League', 'Segunda Division', 'Serie A', 'Serie B',
]

# Markets locked by the Phase 1 / BTTS backtests — accuracy figures are
# the actual backtest output (scripts/backtest_full.py, scripts/backtest_btts.py),
# not estimates.
LOCKED_MARKETS = {
    'dc_1x_ft': {'name': 'DC 1X (FT)', 'accuracy': 72.8, 'weight': 1.0},
    'dc_1x_ht': {'name': 'DC 1X (HT)', 'accuracy': 75.9, 'weight': 1.1},
    'under_3_5': {'name': 'Under 3.5', 'accuracy': 73.1, 'weight': 0.9},
    'gg_ft': {'name': 'BTTS Yes', 'accuracy': 68.6, 'weight': 0.7,
              'high_only': True},  # Only HIGH_SIGNAL
}

# NOTE on suspended markets: a draft spec also asked to cap under_2_5,
# ft_win, under_1_5, and ht_win at TRACKING as "inverted or underperforming".
# Not implemented — under_2_5 actually backtested at +3.3% edge (second-best
# of the four Phase 1 markets), and ft_win/under_1_5/ht_win have never been
# backtested in this codebase at all. Revisit once real backtest data exists
# for those three; don't suspend under_2_5, the data doesn't support it.

MAX_SELECTIONS = 5
MAX_PER_LEAGUE = 2
MIN_EVIDENCE_SCORE = 60  # Minimum composite score to qualify

# ─── BTTS Live Reference ──────────────────────────────────────────────
# A pre-game reference list (not a selection) of fixtures with strong BTTS
# evidence, for use during live in-play matches. Unlike the accumulator
# above, this is NOT restricted to VALIDATED_LEAGUES — it surfaces the
# live pipeline's own gg_ft signal plus each team's actual recent scoring
# record, which is real regardless of whether that league has been
# backtested. What it does NOT do is invent a response-rate percentage
# for leagues nobody has backtested (see BTTS_AWAY_FIRST_RESPONSE below).

# A team must have scored in at least this many of their last
# TEAM_SCORE_LOOKBACK home/away matches to count as "in form" —
# same bar as Track B's "expected to score" gate in scripts/backtest_btts.py.
MIN_TEAM_SCORE_STREAK_FOR_REFERENCE = MODERATE_SIGNAL_MIN
TEAM_SCORE_LOOKBACK = STREAK_WINDOW

# Real Track C output from the full BTTS backtest run
# (scripts/backtest_btts.py, 2026-07-15, btts_report_20260715_1854.xlsx):
# when away scores first in a BTTS-flagged match, home's actual historical
# response rate, per league. Only leagues with samples are listed — a
# league's absence here (including any outside the Phase 1 backtest's
# 18-league set, e.g. MLS) means there is no historical conditional data
# for it, not a 0% rate. The reverse direction (home scores first, does
# away respond?) was never backtested at all — no numbers exist for it.
BTTS_AWAY_FIRST_RESPONSE = {
    'Bundesliga': (67.3, 49),
    'Championship': (61.0, 308),
    'La Liga': (66.7, 45),
    'Ligue 1': (48.1, 27),
    'Premier League': (58.1, 105),
    'Serie A': (63.0, 27),
}


# ─── Data Classes ────────────────────────────────────────────────────

@dataclass
class AccumulatorCandidate:
    """A single potential selection for the accumulator."""
    fixture_id: str
    fixture_date: str
    league_name: str
    home_team: str
    away_team: str
    market_key: str
    market_name: str
    signal_tier: str       # HIGH_SIGNAL or MODERATE_SIGNAL
    alignment_met: bool
    home_venue_streak: int
    home_overall_streak: int
    away_venue_streak: int
    away_overall_streak: int
    evidence_score: float = 0.0
    reasons: List[str] = field(default_factory=list)


@dataclass
class BttsReferenceCard:
    """A single fixture on the BTTS Live Reference List — a fact sheet
    of real recent scoring form, not a prediction."""
    fixture_id: str
    fixture_date: str
    league_name: str
    home_team: str
    away_team: str
    signal_tier: str
    home_scored: int    # e.g. 5 — scored in this many of their last home matches
    home_played: int    # e.g. 5 — out of this many (may be < TEAM_SCORE_LOOKBACK)
    away_scored: int
    away_played: int
    btts_streak: int    # weaker side's BTTS-occurrence streak (0..STREAK_WINDOW)
    btts_window: int


# ─── Evidence Scoring ────────────────────────────────────────────────

def calculate_evidence_score(candidate: AccumulatorCandidate) -> float:
    """
    Composite evidence score (0-100):
    - Streak strength (40%): max streak across all 4 lenses
    - Alignment bonus (15%): +15 if both teams' streaks align
    - League validation (15%): +15 — candidates are pre-filtered to
      VALIDATED_LEAGUES in fetch_candidates, so this is a constant for
      every scored candidate. Kept as its own component so a future
      graduated tier system (once real per-league backtest data exists)
      has somewhere to plug in.
    - Signal tier (15%): HIGH_SIGNAL = +15, MODERATE = +8
    - Market reliability (15%): scaled from backtest accuracy
    """
    score = 0.0
    reasons = []

    max_streak = max(
        candidate.home_venue_streak, candidate.home_overall_streak,
        candidate.away_venue_streak, candidate.away_overall_streak,
    )
    score += min(max_streak / 5 * 40, 40)
    reasons.append(f"Home streak: {candidate.home_venue_streak}v/{candidate.home_overall_streak}o")
    reasons.append(f"Away streak: {candidate.away_venue_streak}v/{candidate.away_overall_streak}o")

    if candidate.alignment_met:
        score += 15
        reasons.append("✅ Both teams' streaks aligned")
    else:
        reasons.append("⚠️ Streaks not fully aligned")

    score += 15
    reasons.append(f"\U0001F4CA {candidate.league_name} (Phase 1 backtest-validated)")

    if candidate.signal_tier == 'HIGH_SIGNAL':
        score += 15
        reasons.append("\U0001F534 HIGH_SIGNAL — maximum streak confidence")
    else:
        score += 8
        reasons.append("\U0001F7E1 MODERATE_SIGNAL")

    market_info = LOCKED_MARKETS.get(candidate.market_key, {})
    market_accuracy = market_info.get('accuracy', 50)
    market_weight = market_info.get('weight', 0.5)
    market_score = (market_accuracy - 50) / 50 * 15 * market_weight
    score += max(0, market_score)
    reasons.append(f"\U0001F4C8 {candidate.market_name} — {market_accuracy}% backtest accuracy")

    candidate.evidence_score = round(score, 1)
    candidate.reasons = reasons
    return score


def is_validated_league(league_name: str) -> bool:
    """Whether a league was covered by the Phase 1 backtest."""
    return league_name in VALIDATED_LEAGUES


# ─── Candidate Selection ────────────────────────────────────────────

def fetch_candidates(cursor) -> List[AccumulatorCandidate]:
    """
    Query flagged_opportunities for today's scan. Filter to locked
    markets and backtest-validated leagues only.
    """
    cursor.execute("""
        SELECT DISTINCT ON (fo.fixture_id, fo.market_key)
            fo.fixture_id, fo.fixture_date, fo.league_name,
            fo.home_team_name, fo.away_team_name,
            fo.market_key, fo.market_name, fo.signal_tier,
            fo.alignment_met,
            fo.home_venue_streak, fo.home_overall_streak,
            fo.away_venue_streak, fo.away_overall_streak
        FROM flagged_opportunities fo
        WHERE fo.scan_date = CURRENT_DATE
          AND fo.signal_tier IN ('HIGH_SIGNAL', 'MODERATE_SIGNAL')
          AND fo.market_key IN %s
        ORDER BY fo.fixture_id, fo.market_key, fo.signal_tier ASC
    """, (tuple(LOCKED_MARKETS.keys()),))

    cols = [d[0] for d in cursor.description]
    rows = [dict(zip(cols, row)) for row in cursor.fetchall()]

    candidates = []
    for row in rows:
        if not is_validated_league(row['league_name']):
            continue  # Skip leagues outside the Phase 1 backtest

        # BTTS only if HIGH_SIGNAL
        market_info = LOCKED_MARKETS.get(row['market_key'], {})
        if market_info.get('high_only') and row['signal_tier'] != 'HIGH_SIGNAL':
            continue

        candidate = AccumulatorCandidate(
            fixture_id=str(row['fixture_id']),
            fixture_date=str(row['fixture_date'] or ''),
            league_name=row['league_name'],
            home_team=row['home_team_name'],
            away_team=row['away_team_name'],
            market_key=row['market_key'],
            market_name=row['market_name'],
            signal_tier=row['signal_tier'],
            alignment_met=row['alignment_met'],
            home_venue_streak=row['home_venue_streak'] or 0,
            home_overall_streak=row['home_overall_streak'] or 0,
            away_venue_streak=row['away_venue_streak'] or 0,
            away_overall_streak=row['away_overall_streak'] or 0,
        )

        calculate_evidence_score(candidate)

        if candidate.evidence_score >= MIN_EVIDENCE_SCORE:
            candidates.append(candidate)

    logger.info(f"Found {len(candidates)} qualifying candidates")
    return candidates


def select_best_5(candidates: List[AccumulatorCandidate]) -> List[AccumulatorCandidate]:
    """
    Select the best MAX_SELECTIONS candidates with diversification constraints:
    - Max 1 selection per fixture
    - Max MAX_PER_LEAGUE selections per league
    - Ranked by evidence score descending
    """
    sorted_candidates = sorted(
        candidates, key=lambda c: c.evidence_score, reverse=True
    )

    selected: List[AccumulatorCandidate] = []
    used_fixtures = set()
    league_counts = defaultdict(int)

    for candidate in sorted_candidates:
        if len(selected) >= MAX_SELECTIONS:
            break

        if candidate.fixture_id in used_fixtures:
            continue

        if league_counts[candidate.league_name] >= MAX_PER_LEAGUE:
            continue

        selected.append(candidate)
        used_fixtures.add(candidate.fixture_id)
        league_counts[candidate.league_name] += 1

    logger.info(f"Selected {len(selected)} for accumulator")
    return selected


# ─── Telegram Formatting ────────────────────────────────────────────

def format_accumulator_message(selections: List[AccumulatorCandidate]) -> str:
    """
    Format the accumulator as a Telegram message.
    Plain text (no MarkdownV2) for reliability.
    """
    if not selections:
        return (
            "\U0001F3AF STREAK ACCUMULATOR — No Selections Today\n\n"
            "No fixtures met the minimum evidence threshold across "
            "locked markets and validated leagues.\n\n"
            "This is the system showing restraint — a tracked feature, "
            "not a failure."
        )

    lines = ["\U0001F3AF STREAK ACCUMULATOR — Daily Best Selections", ""]

    dates = [s.fixture_date[:10] for s in selections if s.fixture_date]
    if dates:
        lines.append(f"\U0001F4C5 Coverage: {min(dates)} to {max(dates)}")
    lines.append(f"\U0001F4CA Selections: {len(selections)}/{MAX_SELECTIONS}")
    lines.append("")
    lines.append("=" * 35)

    for i, sel in enumerate(selections, 1):
        lines.append("")
        lines.append(f"Selection {i}/{len(selections)}:")
        lines.append(f"⚽ {sel.home_team} vs {sel.away_team}")
        lines.append(f"\U0001F3DF️ {sel.league_name}")
        lines.append(f"\U0001F4C5 {sel.fixture_date[:10] if sel.fixture_date else 'TBD'}")
        lines.append(f"\U0001F4CA {sel.market_name}")
        lines.append(f"\U0001F522 Evidence Score: {sel.evidence_score}/100")
        lines.append("")
        lines.append("Why this selection:")
        for reason in sel.reasons:
            lines.append(f"  {reason}")
        lines.append("-" * 35)

    lines += [
        "",
        "⚠️ Evidence-based analysis, not betting advice.",
        "Restraint is a tracked feature.",
        f"Run: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ]

    return "\n".join(lines)


# ─── Excel Report ────────────────────────────────────────────────────

def build_report(selections: List[AccumulatorCandidate]) -> Optional[str]:
    """Single-sheet Excel report of today's selections with full reasoning."""
    from src.spreadsheet_exporter import HAS_OPENPYXL
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed — cannot build accumulator report")
        return None
    if not selections:
        return None

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    BODY_FONT = Font(name='Arial', size=9)

    wb = Workbook()
    ws = wb.active
    ws.title = "Accumulator Selections"
    headers = [
        "#", "Date", "League", "Home", "Away", "Market", "Signal Tier",
        "Aligned", "Home Venue Streak", "Home Overall Streak",
        "Away Venue Streak", "Away Overall Streak", "Evidence Score", "Reasoning",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.freeze_panes = 'A2'

    for i, sel in enumerate(selections, 1):
        row = [
            i, sel.fixture_date[:10] if sel.fixture_date else '', sel.league_name,
            sel.home_team, sel.away_team, sel.market_name, sel.signal_tier,
            'YES' if sel.alignment_met else 'NO',
            sel.home_venue_streak, sel.home_overall_streak,
            sel.away_venue_streak, sel.away_overall_streak,
            sel.evidence_score, " | ".join(sel.reasons),
        ]
        for col_idx, v in enumerate(row, 1):
            ws.cell(row=i + 1, column=col_idx, value=v).font = BODY_FONT

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ''))
             for r in range(1, len(selections) + 2)),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

    filepath = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        f"accumulator_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
    )
    wb.save(filepath)
    logger.info(f"Accumulator report saved: {filepath}")
    return filepath


# ─── BTTS Live Reference ─────────────────────────────────────────────

def team_recent_scored_count(cursor, team_id, venue: str) -> Tuple[int, int]:
    """How many of the team's last TEAM_SCORE_LOOKBACK home/away matches
    did THEY personally score in (regardless of BTTS or the opponent).

    Mirrors scripts/backtest_btts.py's Track B eval_team_scored, but
    queried live against the team's actual most recent matches rather
    than point-in-time. Uses the same completed-status set as the rest
    of the pipeline (FT/AET/PEN), a small generalization from the literal
    'FT'-only query in the original spec — completed matches decided on
    penalties/extra time still have a real final score and shouldn't be
    dropped from a team's scoring form.

    Returns (scored_count, matches_available) — matches_available may be
    less than TEAM_SCORE_LOOKBACK for teams with a short history.
    """
    if venue == 'home':
        cursor.execute("""
            SELECT ft_home > 0 AS scored
            FROM fixtures
            WHERE home_team_id = %s
              AND status IN ('FT', 'AET', 'PEN') AND ft_home IS NOT NULL
            ORDER BY kickoff_utc DESC LIMIT %s
        """, (team_id, TEAM_SCORE_LOOKBACK))
    else:
        cursor.execute("""
            SELECT ft_away > 0 AS scored
            FROM fixtures
            WHERE away_team_id = %s
              AND status IN ('FT', 'AET', 'PEN') AND ft_away IS NOT NULL
            ORDER BY kickoff_utc DESC LIMIT %s
        """, (team_id, TEAM_SCORE_LOOKBACK))
    rows = cursor.fetchall()
    scored = sum(1 for (s,) in rows if s)
    return scored, len(rows)


def fetch_btts_signal_fixtures(cursor) -> List[dict]:
    """Today's gg_ft (BTTS) HIGH/MODERATE signals from the live pipeline scan."""
    cursor.execute("""
        SELECT DISTINCT ON (fo.fixture_id)
            fo.fixture_id, fo.fixture_date, fo.league_name,
            fo.home_team_id, fo.home_team_name,
            fo.away_team_id, fo.away_team_name,
            fo.signal_tier,
            fo.home_venue_streak, fo.home_overall_streak,
            fo.away_venue_streak, fo.away_overall_streak
        FROM flagged_opportunities fo
        WHERE fo.scan_date = CURRENT_DATE
          AND fo.market_key = 'gg_ft'
          AND fo.signal_tier IN ('HIGH_SIGNAL', 'MODERATE_SIGNAL')
        ORDER BY fo.fixture_id, fo.signal_tier ASC
    """)
    cols = [d[0] for d in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def build_btts_live_reference(cursor) -> List[BttsReferenceCard]:
    """
    Build the BTTS Live Reference List: pre-game fixtures where the live
    pipeline already flagged a BTTS (gg_ft) signal AND both teams show
    strong individual scoring form (>= MIN_TEAM_SCORE_STREAK_FOR_REFERENCE
    of their last TEAM_SCORE_LOOKBACK matches), for use as an in-play
    reference during the live match — not a prediction, a fact sheet of
    each team's real recent scoring record.

    Not restricted to VALIDATED_LEAGUES — see the module-level comment
    above BTTS_AWAY_FIRST_RESPONSE for why.
    """
    fixtures = fetch_btts_signal_fixtures(cursor)
    cards: List[BttsReferenceCard] = []

    for fx in fixtures:
        home_scored, home_played = team_recent_scored_count(cursor, fx['home_team_id'], 'home')
        away_scored, away_played = team_recent_scored_count(cursor, fx['away_team_id'], 'away')

        if home_scored < MIN_TEAM_SCORE_STREAK_FOR_REFERENCE or \
           away_scored < MIN_TEAM_SCORE_STREAK_FOR_REFERENCE:
            continue

        home_btts_best = max(fx['home_venue_streak'] or 0, fx['home_overall_streak'] or 0)
        away_btts_best = max(fx['away_venue_streak'] or 0, fx['away_overall_streak'] or 0)
        btts_streak = min(home_btts_best, away_btts_best)  # weaker corroborating side

        cards.append(BttsReferenceCard(
            fixture_id=str(fx['fixture_id']),
            fixture_date=str(fx['fixture_date'] or ''),
            league_name=fx['league_name'],
            home_team=fx['home_team_name'], away_team=fx['away_team_name'],
            signal_tier=fx['signal_tier'],
            home_scored=home_scored, home_played=home_played,
            away_scored=away_scored, away_played=away_played,
            btts_streak=btts_streak, btts_window=STREAK_WINDOW,
        ))

    cards.sort(key=lambda c: (c.fixture_date, c.league_name))
    logger.info(f"BTTS live reference: {len(cards)} qualifying fixtures")
    return cards


def format_btts_reference_message(cards: List[BttsReferenceCard]) -> str:
    """Format the BTTS Live Reference List as a Telegram message.
    Always returns a message (graceful 'no fixtures' text when empty),
    matching format_accumulator_message's precedent."""
    if not cards:
        return (
            "\U0001F4CB BTTS LIVE REFERENCE — No Qualifying Fixtures Today\n\n"
            "No fixtures met the BTTS signal + individual scoring-form bar "
            "for today's live reference list."
        )

    lines = [
        "\U0001F4CB BTTS LIVE REFERENCE — In-Play Betting Guide",
        "",
        "These fixtures have strong BTTS evidence.",
        "Use this list during live matches.",
        "",
    ]

    for c in cards:
        date_str = c.fixture_date[:10] if c.fixture_date else 'TBD'
        time_str = c.fixture_date[11:16] if len(c.fixture_date) > 11 else '??:??'

        lines.append(f"⚽ {c.home_team} vs {c.away_team} ({c.league_name})")
        lines.append(f"\U0001F4C5 {date_str}, {time_str} UTC")
        lines.append(f"\U0001F3E0 {c.home_team} scored in {c.home_scored}/{c.home_played} home matches")
        lines.append(f"✈️ {c.away_team} scored in {c.away_scored}/{c.away_played} away matches")
        lines.append(f"\U0001F4CA BTTS occurred in {c.btts_streak}/{c.btts_window} of both teams' recent matches")
        lines.append("")
        lines.append("\U0001F4CC Live triggers:")

        away_resp = BTTS_AWAY_FIRST_RESPONSE.get(c.league_name)
        if away_resp:
            rate, n = away_resp
            lines.append(
                f"  → IF {c.away_team} scores first: bet BTTS "
                f"({c.league_name} home response rate: {rate}%, {n} samples — "
                f"Phase 1 BTTS backtest)"
            )
        else:
            lines.append(
                f"  → IF {c.away_team} scores first: bet BTTS "
                f"(no backtested response rate for {c.league_name} — "
                f"home scoring form: {c.home_scored}/{c.home_played})"
            )
        lines.append(
            f"  → IF {c.home_team} scores first: bet BTTS "
            f"(away scoring form: {c.away_scored}/{c.away_played} — "
            f"reverse direction has never been backtested)"
        )

        lines.append("")
        lines.append("━" * 24)
        lines.append("")

    lines += [
        "⚠️ Reference only, not betting advice.",
        f"Run: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ]
    return "\n".join(lines)


def build_btts_reference_report(cards: List[BttsReferenceCard]) -> Optional[str]:
    """Single-sheet Excel report of today's BTTS Live Reference List."""
    from src.spreadsheet_exporter import HAS_OPENPYXL
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed — cannot build BTTS reference report")
        return None
    if not cards:
        return None

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    BODY_FONT = Font(name='Arial', size=9)

    wb = Workbook()
    ws = wb.active
    ws.title = "BTTS Live Reference"
    headers = [
        "Date", "League", "Home", "Away", "Signal Tier",
        "Home Scoring (last N)", "Away Scoring (last N)", "BTTS Streak",
        "Away-First Response Rate%", "Away-First Samples",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.freeze_panes = 'A2'

    for i, c in enumerate(cards, 1):
        away_resp = BTTS_AWAY_FIRST_RESPONSE.get(c.league_name)
        row = [
            c.fixture_date[:10] if c.fixture_date else '', c.league_name,
            c.home_team, c.away_team, c.signal_tier,
            f"{c.home_scored}/{c.home_played}", f"{c.away_scored}/{c.away_played}",
            f"{c.btts_streak}/{c.btts_window}",
            away_resp[0] if away_resp else None,
            away_resp[1] if away_resp else None,
        ]
        for col_idx, v in enumerate(row, 1):
            ws.cell(row=i + 1, column=col_idx, value=v).font = BODY_FONT

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ''))
             for r in range(1, len(cards) + 2)),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

    filepath = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        f"btts_live_reference_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
    )
    wb.save(filepath)
    logger.info(f"BTTS live reference report saved: {filepath}")
    return filepath


def run_btts_live_reference() -> dict:
    """
    Independent product: generates + sends the BTTS Live Reference List
    regardless of whether run_accumulator_agent() produced any selections
    (design intent: these are independent products, not conditioned on
    each other). Sent as a separate Telegram message, called after the
    accumulator agent from pipeline.py's Phase 6.

    Respects SHADOW_MODE the same way run_accumulator_agent() does.
    """
    logger.info("=== BTTS Live Reference ===")
    summary = {'fixtures_found': 0, 'telegram_sent': False, 'report_sent': False}

    with get_cursor() as cursor:
        cards = build_btts_live_reference(cursor)
    summary['fixtures_found'] = len(cards)

    msg = format_btts_reference_message(cards)
    filepath = build_btts_reference_report(cards)

    if SHADOW_MODE:
        logger.info(f"SHADOW MODE — BTTS reference not broadcast. Would have sent:\n{msg}")
    else:
        summary['telegram_sent'] = send_telegram_message(msg)
        if filepath:
            from src.spreadsheet_exporter import send_telegram_document
            summary['report_sent'] = send_telegram_document(
                filepath, caption=f"\U0001F4CB BTTS Live Reference — {len(cards)} fixtures"
            )

    logger.info(
        f"BTTS live reference complete: {summary['fixtures_found']} fixtures, "
        f"Telegram: {'sent' if summary['telegram_sent'] else ('shadow' if SHADOW_MODE else 'failed')}"
    )
    return summary


# ─── Main Entry Point ────────────────────────────────────────────────

def run_accumulator_agent() -> dict:
    """
    Main entry: fetch candidates, select best MAX_SELECTIONS, send to
    Telegram.

    Respects SHADOW_MODE the same way telegram_notifier.send_alerts()
    treats detailed HIGH_SIGNAL alerts elsewhere in the pipeline: the
    accumulator is actionable content (a betslip), not a status summary,
    so in shadow mode it's logged but not broadcast. Flip SHADOW_MODE
    off to start sending it live.
    """
    logger.info("=== Accumulator Agent ===")

    summary = {
        'candidates_found': 0,
        'selections_made': 0,
        'telegram_sent': False,
        'report_sent': False,
        'selections': [],
    }

    with get_cursor() as cursor:
        candidates = fetch_candidates(cursor)
        summary['candidates_found'] = len(candidates)
        selections = select_best_5(candidates) if candidates else []

    summary['selections_made'] = len(selections)
    summary['selections'] = [
        {
            'fixture': f"{s.home_team} vs {s.away_team}",
            'market': s.market_name,
            'league': s.league_name,
            'score': s.evidence_score,
        }
        for s in selections
    ]

    for i, sel in enumerate(selections, 1):
        logger.info(
            f"  #{i}: {sel.home_team} vs {sel.away_team} | "
            f"{sel.market_name} | {sel.league_name} | "
            f"Score: {sel.evidence_score}"
        )

    msg = format_accumulator_message(selections)
    filepath = build_report(selections)

    if SHADOW_MODE:
        logger.info(f"SHADOW MODE — accumulator not broadcast. Would have sent:\n{msg}")
    else:
        summary['telegram_sent'] = send_telegram_message(msg)
        if filepath:
            from src.spreadsheet_exporter import send_telegram_document
            summary['report_sent'] = send_telegram_document(
                filepath, caption=f"\U0001F3AF Accumulator report — {len(selections)} selections"
            )

    logger.info(
        f"Accumulator complete: {summary['selections_made']} selections, "
        f"Telegram: {'sent' if summary['telegram_sent'] else ('shadow' if SHADOW_MODE else 'failed')}"
    )
    return summary


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')
    run_accumulator_agent()
