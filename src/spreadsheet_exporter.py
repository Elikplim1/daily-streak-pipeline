"""
Spreadsheet Exporter — generates Excel report and sends via Telegram.

Creates a professional Excel file with all flagged opportunities,
grouped by date and fixture, with clear venue-split streak displays.
Attaches it to Telegram as a document.
"""
import logging
import os
import tempfile
from datetime import datetime
from typing import List, Tuple

import requests

logger = logging.getLogger(__name__)

# Check for openpyxl availability
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed — spreadsheet export disabled")


# Signal tier colors
TIER_FILLS = {
    'HIGH_SIGNAL': PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid') if HAS_OPENPYXL else None,
    'MODERATE_SIGNAL': PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid') if HAS_OPENPYXL else None,
}

HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid') if HAS_OPENPYXL else None
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10) if HAS_OPENPYXL else None
BODY_FONT = Font(name='Arial', size=9) if HAS_OPENPYXL else None


def build_spreadsheet(
    scan_results,
    pipeline_run_id: str,
) -> str:
    """
    Build an Excel spreadsheet from scan results.

    Returns:
        Path to the generated .xlsx file
    """
    if not HAS_OPENPYXL:
        logger.error("Cannot generate spreadsheet — openpyxl not installed")
        return None

    from src.market_presets import CORE_MARKETS

    wb = Workbook()
    ws = wb.active
    ws.title = "Streak Signals"

    # Headers
    headers = [
        "Date", "Kickoff (UTC)", "League", "Home Team", "Away Team",
        "Market", "Signal Tier", "Alignment",
        "Home Venue Streak", "Home Overall", "Home Venue Trend", "Home Overall Trend",
        "Away Venue Streak", "Away Overall", "Away Venue Trend", "Away Overall Trend",
        "Match Type", "Streak Type",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Data rows — only HIGH and MODERATE signals
    row_idx = 2
    for fr in sorted(scan_results, key=lambda x: x.fixture_date or ''):
        for mkey, mdata in fr.market_results.items():
            tier = mdata['signal_tier']
            if tier not in ('HIGH_SIGNAL', 'MODERATE_SIGNAL'):
                continue

            market = CORE_MARKETS.get(mkey)
            if not market:
                continue

            hv = mdata['home_venue']
            ho = mdata['home_overall']
            av = mdata['away_venue']
            ao = mdata['away_overall']

            fixture_date = fr.fixture_date or ''
            date_only = fixture_date[:10] if fixture_date else ''
            time_only = fixture_date[11:16] if len(fixture_date) > 11 else ''

            row_data = [
                date_only, time_only,
                fr.league_name, fr.home_team_name, fr.away_team_name,
                market.name, tier,
                'YES' if mdata['alignment_met'] else 'NO',
                hv.streak_length, ho.streak_length,
                hv.trend_count, ho.trend_count,
                av.streak_length, ao.streak_length,
                av.trend_count, ao.trend_count,
                market.match_type, market.streak_type,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = BODY_FONT

                # Color-code signal tier
                if col_idx == 7:  # Signal Tier column
                    fill = TIER_FILLS.get(tier)
                    if fill:
                        cell.fill = fill
                        cell.font = Font(name='Arial', bold=True, size=9,
                                         color='FFFFFF' if tier == 'HIGH_SIGNAL' else '000000')

            row_idx += 1

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, min(row_idx, 100))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 30)

    # Save to temp file
    filepath = os.path.join(
        tempfile.gettempdir(),
        f"streak_signals_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    wb.save(filepath)
    logger.info(f"Spreadsheet saved: {filepath} ({row_idx - 2} rows)")
    return filepath


def send_telegram_document(filepath: str, caption: str = "") -> bool:
    """Send a file as a Telegram document attachment."""
    from src.config import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID

    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        logger.warning("Telegram credentials not set — skipping document send")
        return False

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"

    try:
        with open(filepath, 'rb') as f:
            resp = requests.post(
                url,
                data={
                    'chat_id': TELEGRAM_CHAT_ID,
                    'caption': caption[:1024],  # Telegram caption limit
                },
                files={'document': (os.path.basename(filepath), f)},
                timeout=60,
            )
        if resp.status_code == 200:
            logger.info("Spreadsheet sent to Telegram")
            return True
        else:
            logger.error(f"Telegram document upload failed: {resp.status_code} {resp.text[:200]}")
            return False
    except Exception as e:
        logger.error(f"Failed to send Telegram document: {e}")
        return False
